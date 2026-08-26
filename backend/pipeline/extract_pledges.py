import sys
sys.stdout.reconfigure(encoding='utf-8')
import json
import re
import openpyxl

from paths import RAW_DIR, INPUT_DIR, OUTPUT_DIR

PATH = RAW_DIR / "00. 민선9기 공약 및 과제분석 종합_개발용.xlsx"
OUT = OUTPUT_DIR / "pledges_data.json"

with open(INPUT_DIR / "competitor_findings.json", encoding="utf-8") as f:
    _raw_findings = json.load(f)
COMPETITOR_FINDINGS = {int(k): v for k, v in _raw_findings.items()}

# Fallback for rows where the B2 sheet's NO column is blank (usually the most
# recently-added pledges), so COMPETITOR_FINDINGS lookup by no can't apply and
# they'd otherwise sit at "미조사" forever. Read directly by the reviewer against
# each project's latest-news text and shaped exactly like a competitor_findings
# entry, so compute_entry_grade()/trust_badge() need no special-casing — these
# just come out tagged "AI 추정 · 출처 미확인" like any other unverified finding.
MANUAL_RESEARCH_OVERRIDES = {
    "인천형 물류 AI 통합 운영시스템 구축": {
        "filing_status": "공고전", "competition_status": None,
        "evidence_summary": "최신동향: '국가전략과제로 격상'—아직 개념·비전 단계, RFP 이전.",
    },
    "AX 가속화를 위한 물류AI 융합원 구축": {
        "filing_status": "공고전", "competition_status": None,
        "evidence_summary": "최신동향: '컨트롤타워 신설...전담기관 통해'—추진체계 설계 이전 단계.",
    },
    "글로벌 AI 오토밸리 조성": {
        "filing_status": "공고전", "competition_status": None,
        "evidence_summary": "최신동향: '사업성 검토 및 개발방향 구체화 예정(용역 통해)'—용역 발주 임박, 선제 기획 적기.",
    },
    "인천 로봇랜드에 피지컬AI 거점 육성": {
        "filing_status": "공고전", "competition_status": None,
        "evidence_summary": "최신동향: '협의체 출범...전략과 과제 도출할 계획'—초기 논의 단계.",
    },
    "로봇 • AI • 반도체 클러스터 조성 및 대기업 유치 (ax와의 관련성 확인 필요)": {
        "filing_status": "공고중", "competition_status": None,
        "evidence_summary": "최신동향: '공모 최종 선정 + 착수보고회 개최'—이미 집행 시작, 공고·추진 국면으로 판단.",
    },
    "수성알파시티의 글로벌 AI • 로봇 하브 성장 방안": {
        "filing_status": "공고중", "competition_status": None,
        "evidence_summary": "최신동향: '국토부 공모 선정 이후 특화단지 조성 본격 추진'—공고·집행 국면으로 판단.",
    },
    "AI의료기기 및 휴머노이드 로봇 관련 대기업 투자 유치": {
        "filing_status": "공고중", "competition_status": None,
        "evidence_summary": "최신동향: '수혜기업 모집공고완료, 8/31까지 상시지원'—명시적 공고문 존재.",
    },
    "AI를 활용한 섬유산업 활성화": {
        "filing_status": "공고중", "competition_status": "이미확정",
        "competitor_name": "한국섬유개발연구원 · 한국섬유소재연구원",
        "evidence_summary": "최신동향: '선정되어...한국섬유개발연구원 & 한국섬유소재연구원'—특정 수행기관 이미 명시.",
    },
}


BUDGET_RE = re.compile(
    r'(\d[\d,]*(?:\.\d+)?\s*조\s*(?:\d[\d,]*(?:\.\d+)?\s*억)?\s*원|\d[\d,]*(?:\.\d+)?\s*억\s*원)'
)
FILING_DATE_RE = re.compile(r'(20\d{2})\s*년.{0,12}?(?:예정|목표|추진)')

GRADE_TO_POTENTIAL = {
    "1순위: 선제 제안 대상": "A",
    "2순위: 공식 입찰 대상": "B",
    "3순위: 경쟁열위": "C",
    "미조사": None,
}


def extract_budget(texts):
    combined = " ".join(t for t in texts if t)
    m = BUDGET_RE.search(combined)
    return m.group(1).replace(" ", "") if m else None


def extract_filing_estimate(texts):
    combined = " ".join(t for t in texts if t)
    m = FILING_DATE_RE.search(combined)
    return f"{m.group(1)}년 예정(추정)" if m else None


BUDGET_PARSE_RE = re.compile(r'^(?:(\d[\d,]*(?:\.\d+)?)조)?(?:(\d[\d,]*(?:\.\d+)?)억)?원?$')


def budget_to_eok(text):
    """Best-effort numeric value (억원 단위) for sorting only — never displayed as-is."""
    if not text:
        return None
    m = BUDGET_PARSE_RE.match(text)
    if not m or not (m.group(1) or m.group(2)):
        return None
    total = 0.0
    if m.group(1):
        total += float(m.group(1).replace(",", "")) * 10000
    if m.group(2):
        total += float(m.group(2).replace(",", ""))
    return total


# A single tracked pledge project realistically tops out in the low-tens-of-조원
# range (e.g. "21조원"). The regex sometimes grabs a bigger figure mentioned in
# the same sentence for surrounding context (a whole mega-cluster, a national
# multi-year total) rather than the project's own budget — e.g. one log entry
# says "...896조 원 규모 호남 반도체 메가클러스터와 연계한..." where 896조 describes
# the cluster, not this AI data-center project. Anything past this threshold is
# more likely a misattributed context figure than a real single-project budget,
# so it's suppressed back to "확인필요" rather than shown as fact.
BUDGET_PLAUSIBLE_MAX_EOK = 300_000  # 30조원


def sanity_check_budget(text, eok):
    if eok is not None and eok > BUDGET_PLAUSIBLE_MAX_EOK:
        return None, None
    return text, eok


def compute_entry_grade(finding):
    if not finding:
        return "미조사"
    filing = finding.get("filing_status")
    comp = finding.get("competition_status")
    if comp == "이미확정":
        return "3순위: 경쟁열위"
    if filing == "공고전":
        return "1순위: 선제 제안 대상"
    if filing == "공고중":
        return "2순위: 공식 입찰 대상"
    return "3순위: 경쟁열위" if comp in ("경쟁사있음(추정)",) else "2순위: 공식 입찰 대상"


wb = openpyxl.load_workbook(PATH, data_only=True, read_only=True)

# ---- C1: region -> 5극3특 grouping + staff ----
ws_c1 = wb["C1"]
rows_c1 = list(ws_c1.iter_rows(values_only=True))[1:]
region_group = {}
staff_by_region = {}
for r in rows_c1:
    if r[0]:
        staff_by_region[r[0]] = r[2]
    # columns 6,7,8 hold: (blank), group name, region list
    if len(r) > 8 and r[7] and r[8]:
        group = r[7]
        for reg in str(r[8]).split(' '):
            region_group[reg.strip()] = group

print("region_group:", region_group)
print("staff_by_region:", staff_by_region)

# ---- B2: main pipeline tracker ----
ws = wb["B2"]
rows = list(ws.iter_rows(values_only=True))
header = rows[1]
idx = {v.strip() if isinstance(v, str) else v: i for i, v in enumerate(header) if v}

# raw log columns: from '최근동향3' (first true log entry after O/P) onward, in (동향,URL) pairs
log_pairs = []
i = 0
while i < len(header):
    h = header[i]
    if h and str(h).startswith('최근동향') and h != '최근동향':
        j = i + 1
        if j < len(header) and header[j] and str(header[j]).startswith('근거URL'):
            log_pairs.append((i, j))
            i = j + 1
            continue
    i += 1

def stage_of(row):
    if row[idx.get('확정', -1)] == 'O':
        return '확정'
    if row[idx.get('유력', -1)] == 'O':
        return '유력'
    if row[idx.get('미정', -1)] == 'O':
        return '미정'
    return '미정'

projects = []
_next_synthetic_no = 900  # fallback id band, well above any real NO value in use
for r in rows[2:]:
    if not r or not r[idx.get('지역', 3)]:
        continue
    region = r[idx['지역']]
    name = r[idx.get('공약명(정제)')]
    no_val = r[idx.get('NO')]
    if not isinstance(no_val, int):
        # blank NO column (usually the most recently-added row) — assign a
        # unique synthetic id so this project never collides with another
        # blank-NO row on id="pd-{no}" or data-jump-no lookups downstream.
        no_val = _next_synthetic_no
        _next_synthetic_no += 1
    finding = COMPETITOR_FINDINGS.get(no_val) or MANUAL_RESEARCH_OVERRIDES.get(name)
    log = []
    for a, b in log_pairs:
        if r[a] or r[b]:
            log.append({"text": r[a], "url": r[b]})
    proj = {
        "no": no_val,
        "region": region,
        "group": region_group.get(region, '기타'),
        "politician": r[idx.get('당선인')],
        "name": name,
        "biz_type": r[idx.get('사업유형')],
        "new_or_existing": r[idx.get('기존/신규')],
        "stage": stage_of(r),
        "org": r[idx.get('주관기관')],
        "ministry": r[idx.get('소관부처')],
        "executor": r[idx.get('수행기관')],
        "latest_news": r[idx.get('최근동향')],
        "latest_url": r[idx.get('근거URL')],
        "anchor": r[idx.get('앵커기업 목록')],
        "grp_partner": r[idx.get('GRP/시니어파트너')],
        "leader": r[idx.get('리더')],
        "assurance": r[idx.get('Assurance')],
        "tax": r[idx.get('Tax')] if 'Tax' in idx else None,
        "deals": r[idx.get('Deals')] if 'Deals' in idx else None,
        "consulting": r[idx.get('Consulting')] if 'Consulting' in idx else None,
        "log": log,
        "staff": staff_by_region.get(region),
        "filing_status": finding.get("filing_status") if finding else None,
        "competition_status": finding.get("competition_status") if finding else None,
        "competitor_name": finding.get("competitor_name") if finding else None,
        "evidence_summary": finding.get("evidence_summary") if finding else None,
        "evidence_url": finding.get("evidence_url") if finding else None,
        "is_yeta": finding.get("is_yeta") if finding else None,
        "entry_grade": compute_entry_grade(finding),
    }
    text_pool = [e.get("text") for e in log] + [proj.get("evidence_summary")]
    proj["budget_text"] = extract_budget(text_pool)
    proj["budget_eok"] = budget_to_eok(proj["budget_text"])
    proj["budget_text"], proj["budget_eok"] = sanity_check_budget(proj["budget_text"], proj["budget_eok"])
    proj["filing_estimate"] = extract_filing_estimate(text_pool)
    proj["potential_grade"] = GRADE_TO_POTENTIAL.get(proj["entry_grade"])
    projects.append(proj)

print(f"총 프로젝트: {len(projects)}")

# ---- A4: national pledge reference (context, linkable by region+politician) ----
ws_a4 = wb["A4"]
rows_a4 = list(ws_a4.iter_rows(values_only=True))
header_a4 = rows_a4[0]
idx_a4 = {v: i for i, v in enumerate(header_a4) if v}
pledges = []
for r in rows_a4[1:]:
    if not r or not r[idx_a4.get('시도', 3)]:
        continue
    pledges.append({
        "sido": r[idx_a4['시도']],
        "politician": r[idx_a4.get('당선인')],
        "party": r[idx_a4.get('정당')],
        "title": r[idx_a4.get('대표공약명')],
        "goal": r[idx_a4.get('세부공약 목표/내용')],
        "field": r[idx_a4.get('공약분야')],
        "detail": r[idx_a4.get('세부내용')],
        "is_ai": r[idx_a4.get('AI_AX관련여부')],
        "ax_level": r[idx_a4.get('AX관련 수준')],
        "keywords": r[idx_a4.get('관련키워드')],
        "idea": r[idx_a4.get('제안아이디어')],
    })
print(f"총 전국 공약: {len(pledges)}")

# ---- C3/C4: PwC anchor account mapping ----
# both sheets: real header is row 0; the "client/group name" column is sparsely filled
# (blank = continuation of the group named in the nearest row above) and needs forward-fill.
def extract_account_sheet(sheet_name, name_col_idx, field_cols):
    """field_cols: dict of {output_key: column_index} for contact fields to pull."""
    ws_ = wb[sheet_name]
    rows_ = list(ws_.iter_rows(values_only=True))
    header_ = rows_[0]
    out = []
    current_group = None
    for r in rows_[1:]:
        if not r or not any(r):
            continue
        name = r[name_col_idx] if name_col_idx < len(r) else None
        if name:
            current_group = str(name).strip()
        if not current_group:
            continue
        rec = {"group": current_group, "row_label": name}
        for key, col in field_cols.items():
            if col < len(r) and r[col]:
                rec[key] = r[col]
        if any(k in rec for k in field_cols):
            out.append(rec)
    return out


# C3 columns (0-indexed): 3=Client명(그룹), 4=세부계열사, 5=시니어파트너, 6=Leader, 7=Assurance, 8=Tax, 9=Deals, 10=Consulting
accounts_c3 = extract_account_sheet("C3", 3, {
    "senior_partner": 5, "leader": 6, "assurance": 7, "tax": 8, "deals": 9, "consulting": 10,
})
# C4 columns: 4=GSP(그룹), 6=FY26 GRP, 7=Global Coordinator, 8=Assurance, 9=Tax, 10=Deals, 12=Consulting, 13=SRP
accounts_c4 = extract_account_sheet("C4", 4, {
    "grp": 6, "global_coordinator": 7, "assurance": 8, "tax": 9, "deals": 10, "consulting": 12, "srp": 13,
})
print(f"C3 계정 매핑: {len(accounts_c3)}, C4 계정 매핑: {len(accounts_c4)}")

payload = {
    "projects": projects,
    "pledges": pledges,
    "accounts_c3": accounts_c3,
    "accounts_c4": accounts_c4,
    "region_group": region_group,
}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(payload, f, ensure_ascii=False, separators=(",", ":"), default=str)

import os
print("크기 MB:", os.path.getsize(OUT) / 1024 / 1024)
